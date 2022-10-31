import openpyxl as xl     #import openpyxl Library    
from openpyxl.chart import BarChart,Reference   
filename= 'marks.xlsx'    #give name of your .xl file to be automated
def process_workbook(filename):     
    wb=xl.load_workbook(filename)
    sheet=wb['Sheet1']

    #cell=sheet['a1']
    #print(cell.value)
    #print(sheet.max_row)

    for row in range(2,sheet.max_row+1):
         cell1=sheet.cell(row,2)    #access values 
         cell2=sheet.cell(row,3)
         cell3=sheet.cell(row,4)
         cell4=sheet.cell(row,5)
         cell=(cell1.value+cell2.value+cell3.value)/3*0.6+10  #logic
         cell_new_column=sheet.cell(row,6)  #to create new column 
         cell_new_column.value=cell
         print(cell)
         
    values=Reference(sheet,         #for bargraph
              min_row=2,
              max_row=sheet.max_row,
              min_col=6,
              max_col=6)
    chart=BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'g2')

    wb.save(filename)
process_workbook(filename)
